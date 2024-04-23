# -*- coding: utf-8 -*-
"""
Created on Tue Apr 16 15:54:02 2024

@author: tokmantsev
"""

import os
import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side

#%% Read the INP and MTX files to create the stiffness matrix

def read_INP(filename):
    
    start_marker = "*Element, type=CAX4"
    end_marker = "*Nset, nset=Set-1, generate"
    capture = False
   
    node_mapping = {}
    number_el =[]
    with open(filename, 'r') as file:
        for line in file:
            line = line.strip() 
            if line == start_marker:
                capture = True  
            elif line == end_marker:
                capture = False 
            elif capture:
                parts = list(map(int, line.split(',')))
                node_mapping[parts[0]-1] = parts[1:5] 
            
                number_el.append(parts[1:5])
        
    number_DOF = max(max(number_el))*2

    return  node_mapping, number_DOF

def read_MTX(filename,t):
    
    matrices = {}
    with open(filename, 'r') as file:
        for line in file:
            element, row, col, value = line.split()
            element, row, col = int(element), int(row), int(col)
            value = float(value)
            
           
            if element not in matrices:
                matrices[element] = np.zeros((t, t))  
            
            
            matrices[element][row-1, col-1] = value
     
            if row != col:
                matrices[element][col-1, row-1] = value

    return matrices

def assemble_global_matrix(element_matrices, node_mapping,dof):

    
    # Create an empty global matrix; size depends on the max node index in node_mapping
    global_matrix = np.zeros(( dof, dof))

    # Assemble each element matrix into the global matrix
    for element, matrix in element_matrices.items():
        nodes = node_mapping[element]
        for i, node_i in enumerate(nodes):
            for j, node_j in enumerate(nodes):
                # Convert node index to global matrix index (x, y)
                global_i_x = 2 * (node_i - 1)    # x-coordinate in the global matrix
                global_i_y = 2 * (node_i - 1) + 1  # y-coordinate in the global matrix
                global_j_x = 2 * (node_j - 1)
                global_j_y = 2 * (node_j - 1) + 1

                # Add local matrix values to the global matrix
                global_matrix[global_i_x, global_j_x] += matrix[2*i, 2*j]
                global_matrix[global_i_x, global_j_y] += matrix[2*i, 2*j+1]
                global_matrix[global_i_y, global_j_x] += matrix[2*i+1, 2*j]
                global_matrix[global_i_y, global_j_y] += matrix[2*i+1, 2*j+1]

    return global_matrix

def generate_labels_notation_abaqus(n):
    """Generate labels based on the number of nodes such as x1, y1, x2, y2, etc."""
    labels = []
    for i in range(1, n//2 + 1):
        labels.extend([f'x{i}', f'y{i}'])
    return labels
def save_matrix_to_excel_notation_abaqus(matrix, filename='output.xlsx'):
    # Check if a file exists with the same name and remove it if it does
    if os.path.exists(filename):
        os.remove(filename)
        print(f"Existing file '{filename}' has been deleted.")

    # Generate labels for the dimensions of the matrix
    labels = generate_labels_notation_abaqus(matrix.shape[1])

    # Convert the NumPy array to a DataFrame with row and column labels
    df = pd.DataFrame(matrix, columns=labels, index=labels)

    # Define border styles

    medium_border = Border(
        right=Side(style='medium'), 
        bottom=Side(style='medium')
    )

    # Use the ExcelWriter to save data with specific formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write the DataFrame to an Excel file with headers
        df.to_excel(writer, header=True, index=True)

        # Access the  worksheet

        worksheet = writer.sheets['Sheet1']

        # Apply formatting to cells
        max_row = df.shape[0] + 1
        max_col = df.shape[1] + 1
        for row_idx in range(2, max_row + 1):  # 1-based index, skip header
            for col_idx in range(2, max_col + 1):  # 1-based index, skip index column
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # Set borders for 2x2 blocks
                if (row_idx - 1) % 2 == 0 and (col_idx - 1) % 2 == 0:
                    # Apply medium border to bottom-right cell of each block
                    cell.border = medium_border

                # Set general styling
                if cell.value == 0 or cell.value == 0.0:
                    cell.number_format = 'General'
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                else:
                    cell.number_format = '0.00E+00'
                if (cell.row - 1) % matrix.shape[1] == (cell.column - 1) % matrix.shape[1]:
                    cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

        # Apply borders to the cells after styling
        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_col):
            for cell in row:
                # Check the position of the cell to decide if it needs a border
                border_right = (cell.column - 1) % 2 == 0 and (cell.column - 1) != max_col - 1
                border_bottom = (cell.row - 1) % 2 == 0 and (cell.row - 1) != max_row - 1
                if border_right and border_bottom:
                    cell.border = Border(
                        right=Side(style='medium'), 
                        bottom=Side(style='medium')
                    )
                elif border_right:
                    cell.border = Border(right=Side(style='medium'))
                elif border_bottom:
                    cell.border = Border(bottom=Side(style='medium'))

    print(f"Matrix saved to '{filename}' with custom formatting.")

def generate_labels_notation_python(n):
    """Generate labels in the order of x1, x2, x3, ..., y1, y2, y3, ..."""
    x_labels = [f'x{i}' for i in range(1, n//2 + 1)]
    y_labels = [f'y{i}' for i in range(1, n//2 + 1)]
    return x_labels + y_labels
def save_matrix_to_excel_notation_python(matrix, filename='output.xlsx'):
    # Check if a file exists with the same name and remove it if it does
    if os.path.exists(filename):
        os.remove(filename)
        print(f"Existing file '{filename}' has been deleted.")

    # Generate labels for the dimensions of the matrix
    labels = generate_labels_notation_python(matrix.shape[1])

    # Convert the NumPy array to a DataFrame with row and column labels
    df = pd.DataFrame(matrix, columns=labels, index=labels)

    # Define border styles

    medium_border = Border(
        right=Side(style='medium'), 
        bottom=Side(style='medium')
    )

    # Use the ExcelWriter to save data with specific formatting
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write the DataFrame to an Excel file with headers
        df.to_excel(writer, header=True, index=True)

        # Access  the worksheet
        worksheet = writer.sheets['Sheet1']

        # Apply formatting to cells
        max_row = df.shape[0] + 1
        max_col = df.shape[1] + 1
        for row_idx in range(2, max_row + 1):  # 1-based index, skip header
            for col_idx in range(2, max_col + 1):  # 1-based index, skip index column
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # Set borders for 2x2 blocks
                if (row_idx - 1) % 2 == 0 and (col_idx - 1) % 2 == 0:
                    # Apply medium border to bottom-right cell of each block
                    cell.border = medium_border

                # Set general styling
                if cell.value == 0 or cell.value == 0.0:
                    cell.number_format = 'General'
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                else:
                    cell.number_format = '0.00E+00'
                if (cell.row - 1) % matrix.shape[1] == (cell.column - 1) % matrix.shape[1]:
                    cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

        # Apply borders to the cells after styling
        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_col):
            for cell in row:
                # Check the position of the cell to decide if it needs a border
                border_right = (cell.column - 1) % 2 == 0 and (cell.column - 1) != max_col - 1
                border_bottom = (cell.row - 1) % 2 == 0 and (cell.row - 1) != max_row - 1
                if border_right and border_bottom:
                    cell.border = Border(
                        right=Side(style='medium'), 
                        bottom=Side(style='medium')
                    )
                elif border_right:
                    cell.border = Border(right=Side(style='medium'))
                elif border_bottom:
                    cell.border = Border(bottom=Side(style='medium'))

    print(f"Matrix saved to '{filename}' with custom formatting.")

#%% Calling of functions to create a Stiffness matrix which follows the Python and Abaqus formats  
repertoire = "fichiers_initiaux"
nom_fichier_INP = "Job-no-BC-4-elements-el-by-el"
nom_fichier_MTX = "Job-no-BC-4-elements-el-by-el"

calling_of_INP = repertoire + "\\" + nom_fichier_INP + ".inp"
calling_of_MTX = repertoire + "\\" + nom_fichier_INP + ".mtx"

dof_per_element = 8 # to modify if needed (8 is number of DOF of each element )

node_mapping, number_of_DOF  = read_INP(calling_of_INP) 

element_matrices = read_MTX(calling_of_MTX, dof_per_element)

K_global_abaqus_format_abaqus = assemble_global_matrix(element_matrices, node_mapping, number_of_DOF)


#%% Rearrangement of the global matrix from Abaqus format (x1,y1,x2,y2...) to Python one (x1,x2... y1,y2...)

def rearrange_matrix_abacus_to_python(matrix):
    n = matrix.shape[0]  # Assume mxm matrix
    assert n % 2 == 0, "The matrix should have an even dimension."

    # Create the new order index
    # First, take all the 'x' indices (even indices)
    # Then, take all the 'y' indices (odd indices)
    new_order = list(range(0, n, 2)) + list(range(1, n, 2))
    
    # Apply the new ordering to the rows and columns
    rearranged_matrix = matrix[:, new_order][new_order, :]
    
    return rearranged_matrix

K_global_abaqus_format_python = rearrange_matrix_abacus_to_python(K_global_abaqus_format_abaqus)




#%% Writing Excel files for each format 
save_matrix_to_excel_notation_abaqus(K_global_abaqus_format_abaqus, 'K_global_abaqus_format_abaqus.xlsx')
save_matrix_to_excel_notation_python(K_global_abaqus_format_python, 'K_global_abaqus_format_python.xlsx')


#%% K matrix directly extracted from python at no boundary condition applied 


mtx_file_path = "fichiers_initiaux\Job-no-BC-2-elements-globale.mtx"
K_global_abaqus_direct_order_abaqus = np.zeros([number_of_DOF, number_of_DOF])


with open(mtx_file_path, 'r') as file:
    for line in file:
        # Each line will be a string. Here's how you might parse it:
        row, col, value = line.split()
        row = int(row)
        col = int(col)
        value = float(value)
        K_global_abaqus_direct_order_abaqus[row-1,col-1] = value